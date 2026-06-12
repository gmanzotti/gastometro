"""
testes/testes_federal.py  —  Testes unitários para pipelines/federal/load.py
──────────────────────────────────────────────────────────────────────────

ESTRUTURA DOS TESTES
  Os testes são organizados por função, na mesma ordem em que aparecem no load.py:

  1. baixar_excel()           → testa download HTTP
  2. _ler_aba_mensal()        → testa leitura e limpeza de aba mensal
  3. _ler_aba_anual()         → testa leitura e limpeza de aba anual
  4. _melt_mensal()           → testa transformação wide → tidy
  5. _computar_pib_por_ano()  → testa cálculo do PIB
  6. _extrair_investimento()  → testa extração das rubricas das abas 1.3/1.3-A
                                (todas as rubricas, com prefixo "INV ")
  7. transformar()            → teste de ponta a ponta com Excel real em memória

COMO RODAR
  Na raiz do projeto:
    pytest testes/testes_federal.py -v

  Para rodar só um grupo:
    pytest testes/testes_federal.py -v -k "pib"

DEPENDÊNCIAS DE TESTE
  - pytest (já no requirements.txt)
  - unittest.mock (stdlib)
  - openpyxl (necessário para criar o Excel mock nos testes de integração;
    é dependência do pandas para leitura de .xlsx — deve estar instalado)
"""

import io
import sys
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest

# Adiciona a raiz do projeto ao path para que o import de pipelines/ funcione
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from pipelines.federal.load import (
    PREFIXO_INVESTIMENTO,
    _computar_pib_por_ano,
    _extrair_investimento,
    _ler_aba_anual,
    _ler_aba_mensal,
    _melt_mensal,
    baixar_excel,
    transformar,
)


# ══════════════════════════════════════════════════════════════════════════
# Fixtures compartilhadas
# ══════════════════════════════════════════════════════════════════════════

# Datas usadas como cabeçalho de coluna nos testes de séries mensais
JAN23 = pd.Timestamp("2023-01-01")
FEV23 = pd.Timestamp("2023-02-01")


def _df_raw_mensal(nome_coluna="Discriminação"):
    """
    Simula o que pd.read_excel retorna ao ler uma aba mensal com header=4.

    Tem 3 linhas:
      - duas de dados numéricos (devem sobreviver ao filtro)
      - uma de nota de rodapé (valor não-numérico na 2ª coluna → deve ser removida)
    """
    return pd.DataFrame({
        nome_coluna: ["1. Receita Total", "1.1 Receita Tributária", "NOTA: Sujeito a revisão"],
        JAN23:       [100.0,              60.0,                      "nota"],
        FEV23:       [110.0,              65.0,                      "nota"],
    })


def _df_raw_anual(nome_coluna="Discriminação"):
    """
    Simula o que pd.read_excel retorna ao ler uma aba anual com header=4.
    Colunas são anos inteiros; inclui uma nota de rodapé.
    """
    return pd.DataFrame({
        nome_coluna: ["1. Receita Total", "NOTA: Sujeito a revisão"],
        2022:        [2000.0,             "nota"],
        2023:        [2200.0,             "nota"],
    })


@pytest.fixture
def excel_rtn_bytes():
    """
    Cria um arquivo Excel mínimo em memória com a estrutura exata da RTN.

    Reproduz as 6 abas que o load.py agora consome:
      1.2, 1.2-A  → séries mensais (nominal e real)
      1.3, 1.3-A  → linha de investimento agregado
      2.2, 2.2-A  → séries anuais para cálculo do PIB

    O cabeçalho institucional (5 linhas) é respeitado para que header=4
    funcione da mesma forma que no Excel real do Tesouro.

    Valores escolhidos para facilitar a verificação manual:
      - Receita mensal: 100 e 110 (jan/fev 2023)
      - Receita anual 2022: 2000, proporção 0.20 → PIB = 10000
      - Receita anual 2023: 2200, proporção 0.22 → PIB = 10000
      - Investimento: 20 e 22 (jan/fev 2023)
    """
    import openpyxl  # importado aqui para falha explícita se não instalado

    wb = openpyxl.Workbook()
    del wb["Sheet"]  # remove a aba padrão criada pelo openpyxl

    def _aba_mensal(nome, receita_jan, receita_fev, label="1. Receita Total", constante=False):
        """Cria uma aba mensal com o cabeçalho de 5 linhas da RTN."""
        ws = wb.create_sheet(nome)
        # Linhas 1-4: cabeçalho institucional (lidas mas descartadas pelo header=4)
        ws.cell(1, 1, "Resultado do Tesouro Nacional")
        ws.cell(2, 1, "Tabela")
        if constante:
            ws.cell(3, 1, "Valores de Mar/2025 - R$ milhões constantes (IPCA)")
        else:
            ws.cell(3, 1, "R$ milhões correntes")
        ws.cell(4, 1, "")
        # Linha 5: cabeçalho das colunas (header=4 no pandas, 0-indexado)
        ws.cell(5, 1, "Discriminação")
        ws.cell(5, 2, datetime(2023, 1, 1))
        ws.cell(5, 3, datetime(2023, 2, 1))
        # Linha 6: primeira linha de dados (iloc[0] no DataFrame resultante)
        ws.cell(6, 1, label)
        ws.cell(6, 2, receita_jan)
        ws.cell(6, 3, receita_fev)
        # Linha 7: nota de rodapé (deve ser filtrada)
        ws.cell(7, 1, "NOTA: Sujeito a revisão")
        ws.cell(7, 2, "nota")
        ws.cell(7, 3, "nota")

    def _aba_anual(nome, receita_2022, receita_2023):
        """Cria uma aba anual com o cabeçalho de 5 linhas da RTN."""
        ws = wb.create_sheet(nome)
        ws.cell(1, 1, "Resultado do Tesouro Nacional")
        ws.cell(2, 1, "Tabela Anual")
        ws.cell(3, 1, "R$ milhões")
        ws.cell(4, 1, "")
        ws.cell(5, 1, "Discriminação")
        ws.cell(5, 2, 2022)
        ws.cell(5, 3, 2023)
        ws.cell(6, 1, "1. Receita Total")
        ws.cell(6, 2, receita_2022)
        ws.cell(6, 3, receita_2023)
        ws.cell(7, 1, "NOTA: Sujeito a revisão")
        ws.cell(7, 2, "nota")
        ws.cell(7, 3, "nota")

    # Séries mensais principais (abas 1.2 e 1.2-A)
    _aba_mensal("1.2",   100.0, 110.0)
    _aba_mensal("1.2-A", 95.0,  104.5, constante=True)

    # Investimento público (abas 1.3 e 1.3-A)
    _aba_mensal("1.3",   20.0, 22.0,  label="Investimento Público")
    _aba_mensal("1.3-A", 19.0, 20.9,  label="Investimento Público", constante=True)

    # Séries anuais para cálculo do PIB (abas 2.2 e 2.2-A)
    _aba_anual("2.2",   2000.0, 2200.0)   # R$ milhões correntes
    _aba_anual("2.2-A", 0.20,   0.22)     # fração decimal do PIB → PIB = 10000

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
# 1. baixar_excel()
# ══════════════════════════════════════════════════════════════════════════

class TestBaixarExcel:
    def test_retorna_bytes_do_conteudo(self):
        """Deve retornar exatamente os bytes recebidos do servidor."""
        conteudo_falso = b"PK\x03\x04fake_excel_bytes"
        mock_resp = MagicMock()
        mock_resp.content = conteudo_falso
        mock_resp.raise_for_status.return_value = None

        with patch("pipelines.federal.load.requests.get", return_value=mock_resp) as mock_get:
            resultado = baixar_excel()

        assert resultado == conteudo_falso
        # Confirma que a URL correta foi chamada com verify=False
        args, kwargs = mock_get.call_args
        assert kwargs.get("verify") is False

    def test_propaga_erro_http(self):
        """Se o servidor retornar erro HTTP, deve lançar exceção."""
        from requests.exceptions import HTTPError

        mock_resp = MagicMock()
        mock_resp.raise_for_status.side_effect = HTTPError("404")

        with patch("pipelines.federal.load.requests.get", return_value=mock_resp):
            with pytest.raises(HTTPError):
                baixar_excel()


# ══════════════════════════════════════════════════════════════════════════
# 2. _ler_aba_mensal()
# ══════════════════════════════════════════════════════════════════════════

class TestLerAbaMensal:
    def test_filtra_notas_de_rodape(self):
        """Linhas onde a 2ª coluna é não-numérica devem ser removidas."""
        raw = _df_raw_mensal()
        mock_xl = MagicMock(spec=pd.ExcelFile)

        with patch("pipelines.federal.load.pd.read_excel", return_value=raw):
            resultado = _ler_aba_mensal(mock_xl, "1.2")

        # raw tem 3 linhas; a nota de rodapé deve ser removida → sobram 2
        assert len(resultado) == 2

    def test_renomeia_primeira_coluna_para_discriminacao(self):
        """A primeira coluna deve sempre se chamar 'discriminacao'."""
        raw = _df_raw_mensal(nome_coluna="Discriminação da Receita")
        mock_xl = MagicMock(spec=pd.ExcelFile)

        with patch("pipelines.federal.load.pd.read_excel", return_value=raw):
            resultado = _ler_aba_mensal(mock_xl, "1.2")

        assert "discriminacao" in resultado.columns

    def test_passa_header_correto_para_read_excel(self):
        """Deve chamar read_excel com header=4 (5ª linha do Excel como cabeçalho)."""
        raw = _df_raw_mensal()
        mock_xl = MagicMock(spec=pd.ExcelFile)

        with patch("pipelines.federal.load.pd.read_excel", return_value=raw) as mock_read:
            _ler_aba_mensal(mock_xl, "1.2")

        _, kwargs = mock_read.call_args
        assert kwargs.get("header") == 4


# ══════════════════════════════════════════════════════════════════════════
# 3. _ler_aba_anual()
# ══════════════════════════════════════════════════════════════════════════

class TestLerAbaAnual:
    def test_filtra_notas_de_rodape(self):
        """Linhas de nota devem ser removidas (mesma lógica da mensal)."""
        raw = _df_raw_anual()
        mock_xl = MagicMock(spec=pd.ExcelFile)

        with patch("pipelines.federal.load.pd.read_excel", return_value=raw):
            resultado = _ler_aba_anual(mock_xl, "2.2")

        # raw tem 2 linhas; a nota deve ser removida → sobra 1
        assert len(resultado) == 1

    def test_renomeia_primeira_coluna_para_discriminacao(self):
        raw = _df_raw_anual(nome_coluna="Rubrica")
        mock_xl = MagicMock(spec=pd.ExcelFile)

        with patch("pipelines.federal.load.pd.read_excel", return_value=raw):
            resultado = _ler_aba_anual(mock_xl, "2.2")

        assert "discriminacao" in resultado.columns


# ══════════════════════════════════════════════════════════════════════════
# 4. _melt_mensal()
# ══════════════════════════════════════════════════════════════════════════

class TestMeltMensal:
    @pytest.fixture
    def df_wide(self):
        """DataFrame wide com 2 indicadores e 2 meses."""
        return pd.DataFrame({
            "discriminacao": ["Receita Total", "Despesa Total"],
            JAN23:           [100.0,           80.0],
            FEV23:           [110.0,           85.0],
        })

    def test_transforma_wide_para_tidy(self, df_wide):
        """Deve retornar uma linha por (indicador, mês)."""
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        # 2 indicadores × 2 meses = 4 linhas
        assert len(resultado) == 4

    def test_colunas_do_resultado(self, df_wide):
        """O resultado deve ter exatamente 3 colunas."""
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        assert set(resultado.columns) == {"discriminacao", "data", "corrente_milhoes"}

    def test_coluna_data_e_datetime(self, df_wide):
        """A coluna 'data' deve ser do tipo datetime (pd.Timestamp)."""
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        assert pd.api.types.is_datetime64_any_dtype(resultado["data"])

    def test_coluna_valor_e_numerica(self, df_wide):
        """A coluna de valor deve ser numérica (float)."""
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        assert pd.api.types.is_numeric_dtype(resultado["corrente_milhoes"])

    def test_strip_em_discriminacao(self):
        """Espaços extras em 'discriminacao' devem ser removidos."""
        df_wide = pd.DataFrame({
            "discriminacao": ["  Receita Total  "],
            JAN23:           [100.0],
        })
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        assert resultado["discriminacao"].iloc[0] == "Receita Total"

    def test_ignora_colunas_nao_timestamp(self):
        """Colunas que não são datas (ex: uma coluna extra de texto) devem ser ignoradas."""
        df_wide = pd.DataFrame({
            "discriminacao": ["Receita"],
            "coluna_extra":  ["texto"],   # não é Timestamp → deve ser ignorada
            JAN23:           [100.0],
        })
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        # Apenas JAN23 é timestamp → 1 linha
        assert len(resultado) == 1

    def test_valores_corretos(self, df_wide):
        """Os valores no resultado devem corresponder aos valores de entrada."""
        resultado = _melt_mensal(df_wide, "corrente_milhoes")

        receita_jan = resultado[
            (resultado["discriminacao"] == "Receita Total") &
            (resultado["data"] == JAN23)
        ]["corrente_milhoes"].iloc[0]

        assert receita_jan == 100.0


# ══════════════════════════════════════════════════════════════════════════
# 5. _computar_pib_por_ano()
# ══════════════════════════════════════════════════════════════════════════

class TestComputarPibPorAno:
    @pytest.fixture
    def df_corr(self):
        """Receita Total em R$ milhões correntes."""
        return pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [2000.0],
            2023:            [2200.0],
        })

    @pytest.fixture
    def df_pib(self):
        """Receita Total como fração decimal do PIB."""
        return pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [0.20],   # 20% do PIB → PIB = 2000/0.20 = 10000
            2023:            [0.22],   # 22% do PIB → PIB = 2200/0.22 = 10000
        })

    def test_calculo_basico(self, df_corr, df_pib):
        """PIB = receita_corrente / fracao_pib."""
        resultado = _computar_pib_por_ano(df_corr, df_pib)

        assert resultado[2022] == pytest.approx(10000.0)
        assert resultado[2023] == pytest.approx(10000.0)

    def test_projeta_ano_seguinte(self, df_corr, df_pib):
        """Deve adicionar uma projeção para o ano seguinte ao último disponível."""
        resultado = _computar_pib_por_ano(df_corr, df_pib)

        # O último ano nos dados é 2023 → deve haver projeção para 2024
        assert 2024 in resultado
        # A projeção usa crescimento de 8%
        assert resultado[2024] == pytest.approx(10000.0 * 1.08)

    def test_nao_projeta_se_proximo_ano_ja_existe(self, df_corr, df_pib):
        """Se 2024 já estiver nos dados, não deve adicionar projeção extra."""
        df_corr_ext = df_corr.copy()
        df_corr_ext[2024] = [2400.0]
        df_pib_ext = df_pib.copy()
        df_pib_ext[2024] = [0.24]

        resultado = _computar_pib_por_ano(df_corr_ext, df_pib_ext)

        # 2024 vem dos dados reais, não da projeção
        assert resultado[2024] == pytest.approx(10000.0)
        # 2025 é a projeção
        assert 2025 in resultado

    def test_ignora_fracao_zero(self, df_corr):
        """Não deve gerar divisão por zero quando a fração do PIB é 0."""
        df_pib_zero = pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [0.0],   # fração zero → deve ser ignorado
        })
        df_corr_single = pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [2000.0],
        })
        resultado = _computar_pib_por_ano(df_corr_single, df_pib_zero)

        # 2022 deve ser omitido (não calcular 0/0 ou 2000/0)
        assert 2022 not in resultado

    def test_retorna_dict_vazio_se_sem_receita_total(self):
        """Se nenhuma linha começar com '1. ', retorna dict vazio."""
        df_sem_receita = pd.DataFrame({
            "discriminacao": ["2. Transferências"],
            2022:            [500.0],
        })
        resultado = _computar_pib_por_ano(df_sem_receita, df_sem_receita)

        assert resultado == {}

    def test_usa_apenas_anos_em_ambas_as_abas(self):
        """
        Só deve calcular PIB para anos presentes nas duas abas (interseção).

        Atenção: o código sempre adiciona uma projeção para o ano seguinte ao
        último calculado. Então se a interseção vai até 2022, 2023 aparece no
        resultado como projeção (pib[2022] * 1.08), não como dado calculado.
        """
        df_corr = pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [2000.0],
            2023:            [2200.0],   # existe no corrente mas não no pib
        })
        df_pib = pd.DataFrame({
            "discriminacao": ["1. Receita Total"],
            2022:            [0.20],     # 2022 é o único ano na interseção
        })
        resultado = _computar_pib_por_ano(df_corr, df_pib)

        # 2022 deve ser calculado a partir dos dados (= 2000/0.20 = 10000)
        assert 2022 in resultado
        assert resultado[2022] == pytest.approx(10000.0)

        # 2023 aparece SOMENTE como projeção (ultimo+1), não como dado calculado
        assert 2023 in resultado
        assert resultado[2023] == pytest.approx(10000.0 * 1.08)

        # 2024 não deve existir (só projeta um ano à frente)
        assert 2024 not in resultado


# ══════════════════════════════════════════════════════════════════════════
# 6. _extrair_investimento()
# ══════════════════════════════════════════════════════════════════════════

class TestExtrairInvestimento:
    @pytest.fixture
    def df_raw_inv_corr(self):
        """
        Aba 1.3 lida do Excel: duas rubricas de dados + uma nota de rodapé.
        Desde 11/06/2026 o pipeline extrai TODAS as rubricas (não só o agregado),
        então as duas primeiras linhas devem sobreviver; a nota deve ser filtrada.
        """
        return pd.DataFrame({
            "Discriminação": ["1. INVESTIMENTO TOTAL", "2.1.1.1 Obras", "NOTA: Sujeito a revisão"],
            JAN23:           [20.0,                    15.0,            "nota"],
            FEV23:           [22.0,                    17.0,            "nota"],
        })

    @pytest.fixture
    def df_raw_inv_cons(self):
        """Aba 1.3-A lida do Excel (mesmas rubricas, R$ constantes)."""
        return pd.DataFrame({
            "Discriminação": ["1. INVESTIMENTO TOTAL", "2.1.1.1 Obras", "NOTA: Sujeito a revisão"],
            JAN23:           [19.0,                    14.0,            "nota"],
            FEV23:           [20.9,                    16.2,            "nota"],
        })

    @pytest.fixture
    def resultado(self, df_raw_inv_corr, df_raw_inv_cons):
        """Executa _extrair_investimento() com as duas abas mockadas."""
        mock_xl = MagicMock(spec=pd.ExcelFile)

        def _fake_read_excel(xl, sheet_name, **kwargs):
            return df_raw_inv_corr if sheet_name == "1.3" else df_raw_inv_cons

        with patch("pipelines.federal.load.pd.read_excel", side_effect=_fake_read_excel):
            return _extrair_investimento(mock_xl)

    def test_extrai_todas_as_rubricas(self, resultado):
        """Todas as rubricas com valor numérico entram (2 rubricas × 2 meses = 4 linhas)."""
        assert len(resultado) == 4
        assert resultado["discriminacao"].nunique() == 2

    def test_filtra_notas_de_rodape(self, resultado):
        """A linha de nota (sem valor numérico) não pode aparecer no resultado."""
        assert not resultado["discriminacao"].str.contains("NOTA").any()

    def test_aplica_prefixo_inv(self, resultado):
        """
        Toda rubrica recebe o prefixo "INV " — evita colisão de numeração com
        a aba 1.2 ("2.1" é FPM/FPE na 1.2, mas Investimentos GND 4 na 1.3).
        """
        assert resultado["discriminacao"].str.startswith(PREFIXO_INVESTIMENTO).all()

    def test_merge_corrente_constante(self, resultado):
        """Cada rubrica/mês deve unir o valor corrente (1.3) ao constante (1.3-A)."""
        obras_jan = resultado[
            (resultado["discriminacao"] == "INV 2.1.1.1 Obras") &
            (resultado["data"] == JAN23)
        ]
        assert len(obras_jan) == 1
        assert obras_jan["corrente_milhoes"].iloc[0] == pytest.approx(15.0)
        assert obras_jan["constante_milhoes"].iloc[0] == pytest.approx(14.0)

    def test_colunas_do_resultado(self, resultado):
        """O resultado deve ter as mesmas colunas que o df principal de transformar()."""
        assert set(resultado.columns) == {"discriminacao", "data", "corrente_milhoes", "constante_milhoes"}


# ══════════════════════════════════════════════════════════════════════════
# 7. transformar()  —  teste de integração com Excel real em memória
# ══════════════════════════════════════════════════════════════════════════

class TestTransformar:
    def test_colunas_do_parquet_de_saida(self, excel_rtn_bytes):
        """O DataFrame final deve ter exatamente as 7 colunas esperadas."""
        df, _ = transformar(excel_rtn_bytes)

        colunas_esperadas = {
            "ano", "mes", "data", "discriminacao",
            "corrente_milhoes", "constante_milhoes", "pct_pib",
        }
        assert set(df.columns) == colunas_esperadas

    def test_inclui_rubricas_de_investimento(self, excel_rtn_bytes):
        """As rubricas das abas 1.3/1.3-A devem entrar no parquet com prefixo 'INV '."""
        df, _ = transformar(excel_rtn_bytes)

        inv = df[df["discriminacao"].str.startswith(PREFIXO_INVESTIMENTO)]
        assert not inv.empty, (
            "Nenhuma rubrica 'INV ' encontrada. "
            "Verifique se _extrair_investimento() foi chamado em transformar()."
        )
        # A rubrica do fixture ("Investimento Público") deve aparecer prefixada
        assert (inv["discriminacao"] == "INV Investimento Público").any()

    def test_inclui_receita_total(self, excel_rtn_bytes):
        """O parquet deve conter a Receita Total (das abas 1.2/1.2-A)."""
        df, _ = transformar(excel_rtn_bytes)

        tem_receita = df["discriminacao"].str.startswith("1. ").any()
        assert tem_receita

    def test_pct_pib_calculado(self, excel_rtn_bytes):
        """
        A coluna pct_pib deve ser não-nula e matematicamente coerente.

        Com os valores do fixture:
          - Receita jan/2023 = 100 R$ milhões
          - PIB anual 2023 = 2200 / 0.22 = 10000 R$ milhões
          - PIB mensal = 10000 / 12 ≈ 833.33 R$ milhões
          - pct_pib = (100 / 833.33) × 100 = 12.0%
        """
        df, _ = transformar(excel_rtn_bytes)

        receita_jan = df[
            (df["discriminacao"] == "1. Receita Total") &
            (df["mes"] == 1)
        ]["pct_pib"]

        assert len(receita_jan) == 1
        assert receita_jan.iloc[0] == pytest.approx(12.0, rel=0.01)

    def test_metadados_tem_base_constante(self, excel_rtn_bytes):
        """O dict de metadados deve conter a chave 'base_constante'."""
        _, meta = transformar(excel_rtn_bytes)

        assert "base_constante" in meta

    def test_metadados_tem_ultima_data(self, excel_rtn_bytes):
        """O dict de metadados deve conter a chave 'ultima_data'."""
        _, meta = transformar(excel_rtn_bytes)

        assert "ultima_data" in meta

    def test_resultado_ordenado(self, excel_rtn_bytes):
        """O DataFrame deve estar ordenado por (discriminacao, ano, mes)."""
        df, _ = transformar(excel_rtn_bytes)

        df_sorted = df.sort_values(["discriminacao", "ano", "mes"]).reset_index(drop=True)
        pd.testing.assert_frame_equal(df, df_sorted)

    def test_sem_colunas_extras(self, excel_rtn_bytes):
        """Não deve haver colunas além das 7 especificadas."""
        df, _ = transformar(excel_rtn_bytes)

        assert len(df.columns) == 7
